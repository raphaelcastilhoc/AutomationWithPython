import openpyxl
from openpyxl.utils import get_column_letter
import Helpers.TextExtractor as TextExtractor

filePath = input('Type the file path: ')

workBook = openpyxl.load_workbook(filePath)
sheetNames = workBook.sheetnames

contacts = []
for sheetName in sheetNames:
    sheet = workBook[sheetName]

    maxColumnLetter = get_column_letter(sheet.max_column)
    maxRow = sheet.max_row

    for rowOfCell in sheet['A1':'%s%s' % (maxColumnLetter, maxRow)]:
        for cell in rowOfCell:
            cellContacts = TextExtractor.extractContacts(cell.value)
            for cellContact in cellContacts:
                contacts.append(cellContact)

print('\n'.join(contacts))